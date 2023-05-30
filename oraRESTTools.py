import json
import requests
import datetime
import xml.etree.ElementTree as ET
import sys, os
import urllib.parse
import csv
import logging
from openpyxl import load_workbook
import xlrd
from urllib.parse import urlsplit
import base64

def getTime():
	currentTime = datetime.datetime.now()
	return currentTime

def setVariables( config ):
	variable = {}
	config = ET.parse(config)
	root = config.getroot()
	for var in root.find('variableList'):
		variable[var.tag] = var.text
	return variable

def setLogging():
	logger = logging.getLogger(__name__)
	logger.setLevel(logging.DEBUG)
	
	fh = logging.FileHandler('psPython.log')
	fh.setLevel(logging.INFO)
	
	ch = logging.StreamHandler()
	ch.setLevel(logging.INFO)
	
	formatter = logging.Formatter('%(asctime)s %(name)s  %(levelname)s \t %(message)s')
	fh.setFormatter(formatter)
	ch.setFormatter(formatter)
	
	logger.addHandler(fh)
	logger.addHandler(ch)
	return logger
		
def getRest( url, session, payload, query, requestHeader, authorization, recordLimit, log, count, *argv ):
	#log = setLogging()
	payload = ''
	
	querystring = { 
					"limit": recordLimit 
					}
	querystring['q'] = query
	
	
	start = getTime()
	urlObject = parseUrl(url)
	
	try:
		r = session.get( url, data=payload, headers=requestHeader, params=querystring, auth=authorization )
		#print ('***', r.status_code, r.text)
		data = r.content
		output = json.loads(data)
		time = getTime() - start
		count += 1
		if argv:
			log.info('\t%s\tStatusCode: %s\t%s sec\t%s' % (count, r.status_code, time, urlObject))
	except:
		output = {'items' : None}
		r.status_code
		time = getTime() - start
		log.info('\t%s\tStatusCode: %s\t**ERROR**%s' %(count, r.status_code, r.text, urlObject))
	
	return output, time, r.status_code, r.text, count

def postRest( url, session, body, requestHeader, authorization, log, count, *argv ):
	#log = setLogging()
	start = getTime()
	urlObject = parseUrl(url)
	#print ("*** MADE IT", url, body)
		
	try:
		r = session.post( url, json=body, headers=requestHeader, auth=authorization )
		#print ( 'XXX', r.status_code, r.text )
		data = r.content
		#print("===",data)
		#print("***", r.content)
		output = json.loads(data)
		time = getTime() - start
		if argv:
			log.info('\t\t%s StatusCode: %s\t%s sec\t%s' % (count, r.status_code, time, urlObject))
		count += 1
	except:
		r.status_code
		time = getTime() - start
		log.info('\t\t%s StatusCode: %s\t**ERROR**%s' %(count, r.status_code, r.text, urlObject))
	
	return output, time, r.status_code, r.text, count

def postBatchRest( url, session, partsList, n, authorization, log, count ):
	#log = setLogging()
	start = getTime()
	batchHeader={'Cache-Control': 'no-cache','Content-Type': 'application/vnd.oracle.adf.batch+json', 'Connection': 'close', 'REST-Framework-Version': "8"}
	urlObject = parseUrl(url)
	
	chunksList = [partsList[i * n:(i + 1) * n] for i in range((len(partsList) + n - 1) // n )] 
	#print ("\n***", chunksList)

	for c in chunksList:
		partsBody = {}
		partsBody['parts'] = c
		#print ("\n", '^^^^',partsBody,'\n\n\n', url, authorization, batchHeader)
		try:
			r = session.post( url, json=partsBody, headers=batchHeader, auth=authorization )
			time = getTime() - start
			log.info('\t\tStatusCode: %s\t%s sec\t%s %s' % (r.status_code, time, urlObject, (count+1)*n))
			#print ('****', r.text, r.status_code)
			count += 1
		except:
			#print ('**** MADE IT')
			r.status_code
			time = getTime() - start
			log.info('\t\tStatusCode: %s\t**ERROR**%s' %(r.status_code, r.text, urlObject))
	
	return time, r.status_code, r.text, count
	
def patchRest( url, session, body, requestHeader, authorization, log, count ):
	#log = setLogging()
	start = getTime()
	urlObject = parseUrl(url)
	
	try:
		r = session.patch( url, json=body, headers=requestHeader, auth=authorization )
		time = getTime() - start
		count += 1
		log.info('\t\t\tStatusCode: %s\t%s sec\t%s' % (r.status_code, time, urlObject))
	except:
		r.status_code
		time = getTime() - start
		count += 1
		log.info('\t\t\tStatusCode: %s\t**ERROR**%s' %(r.status_code, r.text, urlObject))
	
	return time, r.status_code, r.text, count
		
def getResources( filename ):
	resourceNames = []
	with open(filename, 'r', newline = '') as f:  
		for line in f:	
			resourceNames.append(line.rstrip())
	return resourceNames

def getPsPlanId ( psPlanOutput, log ):
	log.info('\tCreating Plan, PlanId cross reference')
	psPlans = []
	psPlanXref = {}
	for p in psPlanOutput['items']:
		psPlans.append( { 
							'PlanId' : p['PlanId'],
							'PlanName': p['PlanName'] 
						} 
						)
		psPlanXref[p['PlanName']] = p['PlanId']
	#log.info('\t\t--> Existing Plans: %s' % ( [dict['PlanName'] for dict in psPlans] ) )
	#log.info('\t\t--> Retrieved Existing Plans')
	return psPlans, psPlanXref
	
def idCode( output, entityKey, entityId, log ):
	#log.info('\t\t %s %s Cross reference' % (entityKey, entityId))
	objectIdCode = {}
	
	for o in output['items']:
		objectIdCode[ o[entityKey] ] = o[ entityId ]
	log.info('\t\t-->Code to Id mapping for\t %s : %s mapping' % ( entityKey, entityId ) )
	return objectIdCode
	
	
def writeCsv ( list, filename, outDir ):
	file = filename + '.csv'
	csvFile = os.path.join( outDir, file)
	with open( csvFile, 'w', newline = '' ) as f:
		header = []
		for h in list[0].keys():
			header.append( h )
		csvwriter= csv.writer(f, delimiter=',', quotechar='"', quoting=csv.QUOTE_MINIMAL)
		csvwriter.writerow( header )
		for i in list:
			w = csv.DictWriter(f, i.keys())
			w.writerow(i)
		f.close()

def getUrl ( *n ):
	params = [i for i in n if i]
	newUrl = '/'.join(params)
	
	return newUrl
	
def getJsonItems ( jsonOutput ):
	objectList = jsonOutput['items']
	
	return objectList

def scmAuth ( user, password ):
	r = requests.Session()
	r.auth = ( user, password )
	#r.headers={	'Cache-Control': 'no-cache','Content-Type': 'application/vnd.oracle.adf.action+json', 'REST-Framework-Version': '8', 'Connection': 'close'}
	r.headers={	'Cache-Control': 'no-cache','Content-Type': 'application/json', 'REST-Framework-Version': '8', 'Connection': 'close'}
	#r.headers = {'Cache-Control': 'no-cache', 'Content-Type': 'application/vnd.oracle.adf.resourceitem+json', 'REST-Framework-Version': '8', 'Connection': 'close'}
	payload = ''
	
	return r, r.auth, r.headers, payload

'''
def readExcel ( filename, object ) :
	wb = load_workbook( filename )
	sheet = wb[ object ]
	res = list( sheet )  				# list of records in excel sheet
	final = []							# List of records as dictionary
	
	for x in range(1, sheet.max_row ):
		partFinal = {}	
		for y in range (0, sheet.max_column):
			# If date, Change from datetime to ISO date format
			if isinstance(res[x][y].value, datetime.datetime):
				partFinal[res[0][y].value] = res[x][y].value.replace(tzinfo=datetime.timezone.utc).isoformat()
			else:
				partFinal[res[0][y].value] = res[x][y].value
		final.append(partFinal)
	
	return final
'''

def parseUrl ( url ):
	parsed = urlsplit(url).path.split('/')[-1]
	
	return parsed
	
def getKey ( links ):
	''' For attribute processing, get the system generated key'''
	for link in links:
		if ( link['rel'] == 'self' ):
			key = parseUrl(link['href'])
	
	return key

def getExcelData ( filename, object ):
	wb = xlrd.open_workbook( filename, formatting_info=True )
	ws = wb.sheet_by_name( object )
	
	col_keys = [ ws.cell(0, col_index).value for col_index in range(ws.ncols) ]	
	excelDictList = []
	
	for row_index in range(1, ws.nrows):
		#d = { col_keys[col_index]: ws.cell(row_index, col_index).value for col_index in range(ws.ncols) }   
		## non-list comprehension version of above
		d={}
		for col_index in range(ws.ncols):
			''' If date - ctype=3 '''
			if ws.cell(row_index, col_index).ctype == 3:
				isoDate = datetime.datetime(*xlrd.xldate_as_tuple(ws.cell(row_index, col_index).value, wb.datemode)).replace(tzinfo=datetime.timezone.utc).isoformat()
				d[ col_keys[col_index] ] = isoDate
			elif (ws.cell(row_index, col_index).ctype == 0) and col_keys[col_index]  != 'Color':
				d[ col_keys[col_index] ] = None
			elif col_keys[col_index] == 'Color':
				''' Get Hex color values for the column with title Color '''
				xfx = ws.cell_xf_index(row_index, col_index)
				xf = wb.xf_list[xfx]
				bgx = xf.background.pattern_colour_index
				pattern_colour = wb.colour_map[bgx]
				hexColour = ('#%02x%02x%02x' % pattern_colour)
				d[ col_keys[col_index] ] = hexColour
			else:
				d[ col_keys[col_index] ] = ws.cell(row_index, col_index).value
				
		excelDictList.append(d)

	return excelDictList

def getParts( id, path, operation, payload):
	''' Get Objects for Batch REST call '''
	parts = {}
	parts['id'] = id
	parts['path'] = path
	parts['operation'] = operation
	parts['payload'] = payload
	
	return parts
	
def getPsBody(action, params):
	body = {}
	body['name'] = action
	if params:
		p = json.loads(params)
		body['parameters'] = [p]

	return body
	
def getEssJobId(output, jobIdField):
	jobId = output[jobIdField]
	
	return jobId

def getLogs(essUrl, jobId, jobDefName, outDir, log, mySession, myAuth, myHeader):
	log.info('\t\t-->Getting ESS Logs %s %s...' % (jobId, jobDefName))

	logName = jobDefName + jobId + ".zip"

	finderUrl = '?finder=ESSJobExecutionDetailsRF;requestId=' + str(jobId) + ',fileType=ALL'
	logsUrl = getUrl(essUrl, finderUrl)

	r = mySession.get( logsUrl, headers=myHeader, params=None, auth=myAuth )
	output = json.loads(r.content)
	items = getEssJobId(output, 'items')
	docContent = base64.b64decode(getEssJobId(items[0], 'DocumentContent'))
	with open(outDir + "/" + logName, "wb") as f:
		f.write(docContent)

def essDetails(essUrl, jobId, log, mySession, myAuth, myHeader):
	log.info('\t\t-->Getting ESS Detailis for %s...' % (jobId))

	finderUrl = '?finder=ESSExecutionDetailsRF;requestId=' + str(jobId)
	detailsUrl = getUrl(essUrl, finderUrl)

	r = mySession.get( detailsUrl, headers=myHeader, params=None, auth=myAuth )
	output = json.loads(r.content)

	items = getEssJobId(output, 'items')
	requestDetailsDict = json.loads(getEssJobId(items[0], 'RequestStatus'))
	essDetails = json.dumps(requestDetailsDict, sort_keys=False, indent=10)

	log.info('%s' % (essDetails))

def toBase64(inputDir, file):
	zipFile = os.path.join(inputDir, file)
	with open(zipFile,'rb') as f:
		f_data = f.read()
		base64_encoded = base64.b64encode(f_data)
		base64_message = base64_encoded.decode('utf-8')

	return base64_message


#TODO, zip archive (using shutil) shutil.make_archive('C:\\Users\\alkim.ORADEV\\Documents\\1PS\\fbdi\\1\\WisWdImport', 'zip', 'C:\\Users\\alkim.ORADEV\\Documents\\1PS\\fbdi\\1\\WisWdImport')
#TODO write ess job status for parent child
